#!/usr/bin/env python3
"""
Generate CSV Reports from JSON Data
This script generates CSV reports from the JSON data produced by the parsers.
It can generate reports for specific students, modules, or the entire class.
"""

import os
import json
import csv
import argparse
import re
from typing import Dict, List, Any, Optional, Tuple

def load_json_data(file_path: str) -> Dict[str, Any]:
    """
    Load JSON data from a file.
    
    Args:
        file_path: Path to the JSON file
        
    Returns:
        Dictionary containing the JSON data
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def extract_module_number(resource_name: str) -> Optional[int]:
    """
    Extract the module number from a resource name.
    
    Args:
        resource_name: Name of the resource
        
    Returns:
        Module number as an integer or None if not found
    """
    # Look for patterns like "1.1.7", "2.3.4", etc.
    match = re.search(r'(\d+)\.', resource_name)
    if match:
        return int(match.group(1))
    return None

def get_module_resources(resources: List[Dict[str, Any]], module_number: int, resource_type: str = None) -> List[Dict[str, Any]]:
    """
    Get all resources for a specific module and optionally filter by resource type.
    
    Args:
        resources: List of resources
        module_number: Module number to filter by
        resource_type: Type of resource to filter by (e.g., 'lab', 'fact_sheet', 'assessment')
                      If None, returns all resources for the module
        
    Returns:
        List of resources for the specified module and resource type
    """
    module_resources = [r for r in resources if extract_module_number(r['name']) == module_number]
    
    if resource_type:
        if resource_type == 'lab':
            return [r for r in module_resources if r['name'].startswith('Lab')]
        elif resource_type == 'fact_sheet':
            return [r for r in module_resources if r['name'].startswith('Fact Sheet')]
        elif resource_type == 'assessment':
            return [r for r in module_resources if r['name'].startswith('Assessment')]
        else:
            return []
    
    return module_resources

def count_completed_resources(resources: List[Dict[str, Any]]) -> int:
    """
    Count the number of completed resources.
    
    Args:
        resources: List of resources
        
    Returns:
        Number of completed resources
    """
    return sum(1 for r in resources if r['completion'] == 1.0)

def get_total_time_for_resources(resource_time_data: Dict[str, Any], 
                                 student_name: str, 
                                 resource_type: str, 
                                 module_number: int) -> int:
    """
    Get the total time spent on resources of a specific type for a specific module.
    
    Args:
        resource_time_data: Dictionary containing resource time data
        student_name: Name of the student
        resource_type: Type of resource (lab, assessment, fact_sheet)
        module_number: Module number to filter by
        
    Returns:
        Total time spent in seconds
    """
    # Find the student in the resource time data
    student_data = None
    for student in resource_time_data['students']:
        if student['name'] == student_name:
            student_data = student
            break
    
    if not student_data:
        return 0
    
    # Get resources of the specified type for the specified module
    resources = student_data['resources'].get(resource_type, [])
    module_resources = [r for r in resources if extract_module_number(r['name']) == module_number]
    
    # Calculate total time
    total_time = sum(r['time_seconds'] for r in module_resources)
    
    return total_time

def format_time(seconds: int) -> str:
    """
    Format seconds to a time string in the format "Xh Ym Zs".
    
    Args:
        seconds: Total seconds
        
    Returns:
        Formatted time string
    """
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    
    result = ""
    if hours > 0:
        result += f"{hours}h "
    if minutes > 0 or hours > 0:
        result += f"{minutes}m "
    result += f"{secs}s"
    
    return result.strip()

def calculate_modules_behind(student_data: Dict[str, Any], current_module: int) -> int:
    """
    Calculate how many modules a student is behind.
    
    Args:
        student_data: Dictionary containing student data
        current_module: Current module being taught
        
    Returns:
        Number of modules behind
    """
    # Get all completed modules
    completed_modules = set()
    
    # Check labs
    for lab in student_data['labs']:
        module = extract_module_number(lab['name'])
        if module and lab['completion'] == 1.0:
            completed_modules.add(module)
    
    # Count modules behind (modules not completed up to current module)
    modules_behind = 0
    for module in range(1, current_module + 1):
        if module not in completed_modules:
            modules_behind += 1
    
    return modules_behind

def get_average_assessment_score(student_data: Dict[str, Any], module_number: int) -> float:
    """
    Get the average assessment score for a specific module.
    
    Args:
        student_data: Dictionary containing student data
        module_number: Module number to filter by
        
    Returns:
        Average assessment score as a percentage
    """
    module_assessments = [a for a in student_data['assessments'] 
                         if extract_module_number(a['name']) == module_number]
    
    if not module_assessments:
        return 0.0
    
    total_score = sum(a['completion'] for a in module_assessments)
    return (total_score / len(module_assessments)) * 100

def generate_module_report(assessment_data_path: str, 
                          resource_time_data_path: str, 
                          output_path: str, 
                          module_number: int,
                          student_name: Optional[str] = None,
                          current_module: int = None) -> None:
    """
    Generate a report for a specific module.
    
    Args:
        assessment_data_path: Path to the assessment data JSON file
        resource_time_data_path: Path to the resource time data JSON file
        output_path: Path to save the CSV report
        module_number: Module number to generate the report for
        student_name: Name of the student to filter by (optional)
        current_module: Current module being taught (used for calculating modules behind)
    """
    # Use module_number as current_module if not specified
    if current_module is None:
        current_module = module_number
        
    # Load data
    assessment_data = load_json_data(assessment_data_path)
    resource_time_data = load_json_data(resource_time_data_path)
    
    # Filter students if needed
    students = assessment_data['students']
    if student_name:
        students = [s for s in students if s['name'] == student_name]
        if not students:
            print(f"No student found with name: {student_name}")
            return
    
    # Prepare CSV data
    csv_data = []
    
    for student in students:
        # Get module labs
        module_labs = get_module_resources(student['labs'], module_number, 'lab')
        total_module_labs = len(module_labs)
        completed_module_labs = count_completed_resources(module_labs)
        
        # Calculate labs behind for this module
        labs_behind = total_module_labs - completed_module_labs
        
        # Get time spent on module labs
        lab_time_seconds = get_total_time_for_resources(
            resource_time_data, student['name'], 'lab', module_number
        )
        
        # Get assessment scores
        assessment_score = get_average_assessment_score(student, module_number)
        
        # Add to CSV data
        csv_data.append({
            'Student Name': student['name'],
            'Email': student['email'],
            f'Module {module_number} Labs Completed': completed_module_labs,
            f'Total Module {module_number} Labs': total_module_labs,
            f'Completion Percentage': f"{(completed_module_labs / total_module_labs * 100) if total_module_labs > 0 else 0:.1f}%",
            f'Time Spent on Module {module_number} Labs': format_time(lab_time_seconds),
            f'Time Spent (seconds)': lab_time_seconds,
            f'Module {module_number} Assessment Score': f"{assessment_score:.1f}%",
            'Labs Behind': labs_behind
        })
    
    # Write to CSV
    if csv_data:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
            writer.writeheader()
            writer.writerows(csv_data)
        
        print(f"Module {module_number} report saved to {output_path}")
    else:
        print(f"No data found for module {module_number}")

def generate_student_report(assessment_data_path: str, 
                           resource_time_data_path: str, 
                           output_path: str, 
                           student_name: str) -> None:
    """
    Generate a report for a specific student across all modules.
    
    Args:
        assessment_data_path: Path to the assessment data JSON file
        resource_time_data_path: Path to the resource time data JSON file
        output_path: Path to save the CSV report
        student_name: Name of the student to generate the report for
    """
    # Load data
    assessment_data = load_json_data(assessment_data_path)
    resource_time_data = load_json_data(resource_time_data_path)
    
    # Find the student
    student_data = None
    for student in assessment_data['students']:
        if student['name'] == student_name:
            student_data = student
            break
    
    if not student_data:
        print(f"Student '{student_name}' not found")
        return
    
    # Find all modules
    all_modules = set()
    for resource in student_data['labs'] + student_data['assessments']:
        module = extract_module_number(resource['name'])
        if module:
            all_modules.add(module)
    
    # Prepare CSV data
    csv_data = []
    
    for module in sorted(all_modules):
        # Get module labs
        module_labs = get_module_resources(student_data['labs'], module, 'lab')
        total_module_labs = len(module_labs)
        completed_module_labs = count_completed_resources(module_labs)
        
        # Get time spent on module labs
        lab_time_seconds = get_total_time_for_resources(
            resource_time_data, student_data['name'], 'lab', module
        )
        
        # Get assessment scores
        assessment_score = get_average_assessment_score(student_data, module)
        
        # Calculate modules behind
        modules_behind = calculate_modules_behind(student_data, module)
        
        # Add to CSV data
        csv_data.append({
            'Module': module,
            'Labs Completed': completed_module_labs,
            'Total Labs': total_module_labs,
            'Completion Percentage': f"{(completed_module_labs / total_module_labs * 100) if total_module_labs > 0 else 0:.1f}%",
            'Time Spent on Labs': format_time(lab_time_seconds),
            'Time Spent (seconds)': lab_time_seconds,
            'Assessment Score': f"{assessment_score:.1f}%",
            'Modules Behind': modules_behind
        })
    
    # Write to CSV
    if csv_data:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
            writer.writeheader()
            writer.writerows(csv_data)
        
        print(f"Student report for {student_name} saved to {output_path}")
    else:
        print(f"No data found for student {student_name}")

def generate_class_report(assessment_data_path: str, 
                         resource_time_data_path: str, 
                         output_path: str,
                         current_module: int) -> None:
    """
    Generate a report for the entire class.
    
    Args:
        assessment_data_path: Path to the assessment data JSON file
        resource_time_data_path: Path to the resource time data JSON file
        output_path: Path to save the CSV report
        current_module: Current module being taught
    """
    # Load data
    assessment_data = load_json_data(assessment_data_path)
    resource_time_data = load_json_data(resource_time_data_path)
    
    # Find all modules in the data
    all_modules = set()
    for student in assessment_data['students']:
        for lab in student['labs']:
            module_number = extract_module_number(lab['name'])
            if module_number:
                all_modules.add(module_number)
    
    # Prepare CSV data
    csv_data = []
    
    for student in assessment_data['students']:
        # Calculate total labs and completed labs
        total_labs = len(student['labs'])
        completed_labs = count_completed_resources(student['labs'])
        
        # Calculate completion percentage
        completion_percentage = (completed_labs / total_labs) * 100 if total_labs > 0 else 0
        
        # Calculate total time spent on labs
        total_lab_time = 0
        for module in all_modules:
            total_lab_time += get_total_time_for_resources(
                resource_time_data, student['name'], 'lab', module
            )
        
        # Calculate average assessment score
        total_score = sum(a['completion'] for a in student['assessments'])
        avg_assessment_score = (total_score / len(student['assessments'])) * 100 if student['assessments'] else 0
        
        # Calculate labs behind
        labs_behind = total_labs - completed_labs
        
        # Add to CSV data
        csv_data.append({
            'Student Name': student['name'],
            'Email': student['email'],
            'Total Labs Completed': completed_labs,
            'Total Labs': total_labs,
            'Completion Percentage': f"{completion_percentage:.1f}%",
            'Total Time Spent on Labs': format_time(total_lab_time),
            'Time Spent (seconds)': total_lab_time,
            'Average Assessment Score': f"{avg_assessment_score:.1f}%",
            'Labs Behind': labs_behind
        })
    
    # Sort by completion percentage (descending)
    csv_data.sort(key=lambda x: float(x['Completion Percentage'].strip('%')), reverse=True)
    
    # Write to CSV
    with open(output_path, 'w', newline='') as csvfile:
        fieldnames = [
            'Student Name', 'Email', 'Total Labs Completed', 'Total Labs', 
            'Completion Percentage', 'Total Time Spent on Labs', 'Time Spent (seconds)',
            'Average Assessment Score', 'Labs Behind'
        ]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for row in csv_data:
            writer.writerow(row)
    
    print(f"Class report saved to {output_path}")

def generate_excel_report(assessment_data_path: str,
                        resource_time_data_path: str,
                        output_path: str,
                        current_module: int) -> None:
    """
    Generate an Excel report with multiple sheets - one for each module plus a class summary.
    
    Args:
        assessment_data_path: Path to the assessment data JSON file
        resource_time_data_path: Path to the resource time data JSON file
        output_path: Path to save the Excel report
        current_module: Current module being taught
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("Required packages not found. Installing pandas and openpyxl...")
        import subprocess
        subprocess.check_call(["uv", "pip", "install", "pandas", "openpyxl"])
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill
    
    # Create a new workbook
    wb = Workbook()
    
    # Generate class report data
    assessment_data = load_json_data(assessment_data_path)
    resource_time_data = load_json_data(resource_time_data_path)
    
    # Find all modules in the data
    all_modules = set()
    for student in assessment_data['students']:
        for lab in student['labs']:
            module_number = extract_module_number(lab['name'])
            if module_number:
                all_modules.add(module_number)
    
    # Create class summary sheet
    ws_summary = wb.active
    ws_summary.title = "Class Summary"
    
    # Prepare class summary data
    class_data = []
    
    for student in assessment_data['students']:
        # Calculate total labs and completed labs across all modules
        total_labs = len(student['labs'])
        completed_labs = count_completed_resources(student['labs'])
        
        # Calculate total time spent on labs
        total_lab_time = 0
        for module in all_modules:
            total_lab_time += get_total_time_for_resources(
                resource_time_data, student['name'], 'lab', module
            )
        
        # Calculate average assessment score across all modules
        total_score = sum(a['completion'] for a in student['assessments'])
        avg_assessment_score = (total_score / len(student['assessments'])) * 100 if student['assessments'] else 0
        
        # Calculate labs behind
        labs_behind = total_labs - completed_labs
        
        # Add to data
        class_data.append({
            'Student Name': student['name'],
            'Email': student['email'],
            'Total Labs Completed': completed_labs,
            'Total Labs': total_labs,
            'Completion Percentage': f"{(completed_labs / total_labs * 100) if total_labs > 0 else 0:.1f}%",
            'Total Time Spent on Labs': format_time(total_lab_time),
            'Average Assessment Score': f"{avg_assessment_score:.1f}%",
            'Labs Behind': labs_behind
        })
    
    # Sort by completion percentage (descending)
    class_data.sort(key=lambda x: float(x['Completion Percentage'].strip('%')), reverse=True)
    
    # Convert to DataFrame for easier Excel handling
    df_summary = pd.DataFrame(class_data)
    
    # Add to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True)):
        for c_idx, value in enumerate(row):
            cell = ws_summary.cell(row=r_idx+1, column=c_idx+1, value=value)
            # Format header row
            if r_idx == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Create module sheets
    for module in sorted(all_modules):
        # Create new sheet
        ws = wb.create_sheet(title=f"Module {module}")
        
        # Generate module report data
        module_data = []
        
        for student in assessment_data['students']:
            # Get module labs
            module_labs = get_module_resources(student['labs'], module, 'lab')
            total_module_labs = len(module_labs)
            completed_module_labs = count_completed_resources(module_labs)
            
            # Skip if student has no labs for this module
            if total_module_labs == 0:
                continue
            
            # Calculate labs behind for this module
            labs_behind = total_module_labs - completed_module_labs
            
            # Get time spent on module labs
            lab_time_seconds = get_total_time_for_resources(
                resource_time_data, student['name'], 'lab', module
            )
            
            # Get assessment scores
            assessment_score = get_average_assessment_score(student, module)
            
            # Add to data
            module_data.append({
                'Student Name': student['name'],
                'Email': student['email'],
                f'Module {module} Labs Completed': completed_module_labs,
                f'Total Module {module} Labs': total_module_labs,
                'Completion Percentage': f"{(completed_module_labs / total_module_labs * 100) if total_module_labs > 0 else 0:.1f}%",
                f'Time Spent on Module {module} Labs': format_time(lab_time_seconds),
                f'Module {module} Assessment Score': f"{assessment_score:.1f}%",
                'Labs Behind': labs_behind
            })
        
        # Sort by completion percentage (descending)
        if module_data:
            module_data.sort(key=lambda x: float(x['Completion Percentage'].strip('%')), reverse=True)
            
            # Convert to DataFrame for easier Excel handling
            df_module = pd.DataFrame(module_data)
            
            # Add to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df_module, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx+1, column=c_idx+1, value=value)
                    # Format header row
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel report with multiple sheets saved to {output_path}")

def main():
    """Main function to parse command line arguments and execute the report generation."""
    parser = argparse.ArgumentParser(description='Generate CSV reports from JSON data.')
    parser.add_argument('--assessment-data', 
                        help='Path to the assessment data JSON file')
    parser.add_argument('--resource-time-data', 
                        help='Path to the resource time data JSON file')
    parser.add_argument('--output', 
                        help='Path to save the CSV report')
    parser.add_argument('--module', type=int, help='Module number to generate the report for')
    parser.add_argument('--student', help='Student name to generate the report for')
    parser.add_argument('--current-module', type=int, default=7,
                        help='Current module being taught (used for calculating modules behind)')
    parser.add_argument('--excel', action='store_true', help='Generate an Excel report instead of a CSV report')
    parser.add_argument('--date', help='Date folder to use (format: YY-MM-DD)')
    parser.add_argument('--all', action='store_true',
                        help='Generate all reports: module reports, student reports, class report, and Excel report')
    
    args = parser.parse_args()
    
    # Handle date-based directory structure
    if args.date:
        date_folder = args.date
    else:
        # Find the latest date folder
        processed_dirs = [d for d in os.listdir(os.path.join("assets", "processed")) 
                         if os.path.isdir(os.path.join("assets", "processed", d))]
        if not processed_dirs:
            print("No processed data folders found.")
            return
        date_folder = sorted(processed_dirs)[-1]  # Get the latest date
        print(f"Using latest date folder: {date_folder}")
    
    processed_date_dir = os.path.join("assets", "processed", date_folder)
    reports_date_dir = os.path.join("reports", date_folder)
    
    # Ensure reports directory exists
    os.makedirs(reports_date_dir, exist_ok=True)
    
    # Find the assessment data file
    if args.assessment_data:
        assessment_data_path = args.assessment_data
    else:
        # Find the latest classgradebook file in the date folder
        assessment_files = [f for f in os.listdir(processed_date_dir) 
                          if f.startswith("classgradebook") and f.endswith(".json")]
        if not assessment_files:
            print(f"No assessment data files found in {processed_date_dir}")
            return
        assessment_data_path = os.path.join(processed_date_dir, assessment_files[0])
        print(f"Using assessment data: {assessment_data_path}")
    
    # Find the resource time data file
    if args.resource_time_data:
        resource_time_data_path = args.resource_time_data
    else:
        # Find the latest timeperresource file in the date folder
        resource_time_files = [f for f in os.listdir(processed_date_dir) 
                             if f.startswith("timeperresource") and f.endswith(".json")]
        if not resource_time_files:
            print(f"No resource time data files found in {processed_date_dir}")
            return
        resource_time_data_path = os.path.join(processed_date_dir, resource_time_files[0])
        print(f"Using resource time data: {resource_time_data_path}")
    
    # If --all flag is set, generate all reports
    if args.all:
        print("Generating all reports...")
        
        # Load assessment data to get list of students and modules
        assessment_data = load_json_data(assessment_data_path)
        
        # Get list of all modules
        all_modules = set()
        for student in assessment_data['students']:
            for lab in student['labs']:
                module_number = extract_module_number(lab['name'])
                if module_number:
                    all_modules.add(module_number)
        
        # Generate class report
        class_report_path = os.path.join(reports_date_dir, "class_report.csv")
        generate_class_report(
            assessment_data_path, 
            resource_time_data_path, 
            class_report_path,
            args.current_module
        )
        print(f"Class report saved to {class_report_path}")
        
        # Generate module reports for all modules
        for module in sorted(all_modules):
            module_report_path = os.path.join(reports_date_dir, f"module_{module}_report.csv")
            generate_module_report(
                assessment_data_path,
                resource_time_data_path,
                module_report_path,
                module,
                current_module=args.current_module
            )
            print(f"Module {module} report saved to {module_report_path}")
        
        # Generate student reports for all students
        for student in assessment_data['students']:
            student_name = student['name']
            student_report_path = os.path.join(reports_date_dir, f"student_{student_name.replace(', ', '_').replace(' ', '_')}_report.csv")
            generate_student_report(
                assessment_data_path,
                resource_time_data_path,
                student_report_path,
                student_name
            )
            print(f"Student report for {student_name} saved to {student_report_path}")
        
        # Generate Excel report
        excel_report_path = os.path.join(reports_date_dir, "class_report.xlsx")
        generate_excel_report(
            assessment_data_path,
            resource_time_data_path,
            excel_report_path,
            args.current_module
        )
        print(f"Excel report with multiple sheets saved to {excel_report_path}")
        
        return
    
    # Set default output path if not specified
    if not args.output:
        # Extract base names for report naming
        assessment_base = os.path.splitext(os.path.basename(assessment_data_path))[0]
        resource_base = os.path.splitext(os.path.basename(resource_time_data_path))[0]
        
        if args.module and args.student:
            report_name = f"module{args.module}_{args.student.replace(' ', '_')}_report"
        elif args.module:
            report_name = f"module{args.module}_report"
        elif args.student:
            report_name = f"{args.student.replace(' ', '_')}_report"
        else:
            report_name = "class_report"
        
        if args.excel:
            output_path = os.path.join(reports_date_dir, f"{report_name}.xlsx")
        else:
            output_path = os.path.join(reports_date_dir, f"{report_name}.csv")
    else:
        output_path = args.output
    
    print(f"Using assessment data: {assessment_data_path}")
    print(f"Using resource time data: {resource_time_data_path}")
    print(f"Output will be saved to: {output_path}")
    
    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Handle specific report generation
    if args.module and args.student:
        # Generate report for a specific student and module
        if args.excel:
            generate_excel_report(
                assessment_data_path,
                resource_time_data_path,
                output_path,
                args.current_module
            )
        else:
            generate_module_report(
                assessment_data_path, 
                resource_time_data_path, 
                output_path, 
                args.module,
                args.student,
                args.current_module
            )
    elif args.module:
        # Generate module report
        if args.excel:
            generate_excel_report(
                assessment_data_path,
                resource_time_data_path,
                output_path,
                args.current_module
            )
        else:
            generate_module_report(
                assessment_data_path, 
                resource_time_data_path, 
                output_path, 
                args.module,
                current_module=args.current_module
            )
    elif args.student:
        # Generate student report
        generate_student_report(
            assessment_data_path, 
            resource_time_data_path, 
            output_path, 
            args.student
        )
    else:
        # Generate class report or Excel report
        if args.excel:
            generate_excel_report(
                assessment_data_path,
                resource_time_data_path,
                output_path,
                args.current_module
            )
        else:
            generate_class_report(
                assessment_data_path, 
                resource_time_data_path, 
                output_path,
                args.current_module
            )

if __name__ == "__main__":
    main()
